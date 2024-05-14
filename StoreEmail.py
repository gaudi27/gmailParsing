#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Apr 15 18:07:25 2024

@author: George Audi
"""

"""this is a program to save emails to a txt file so that the emails are saved
and stops repeats from printing in the excel sheet"""

#use openpyxl to read the emails for repeats
import openpyxl
import os


#checks to see if file exists, if not it creates a new file for the emails to be pasted into
def excel_file_exists(file_path):
    if os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(filename=file_path)
            wb.close()  # Close the workbook if it's successfully opened
            return True
        except:
            return False
    else:
        return False


#checks for if a new file needs creation
def needNewFile():
    file_path = "List of Emails to be used.xlsx"
    if excel_file_exists(file_path) == False:
        wb = openpyxl.Workbook()
        ws = wb.active
        x = ["Sender", " ","Subject", "Body"]
        ws.append(x)
        ws.title = "List of Emails to be used"
        wb.save(filename=file_path)


def EmailStorage(dic):
    sheet = openpyxl.load_workbook("List of Emails to be used.xlsx")
    sheetData = sheet.active
    maxRow = sheetData.max_row
    bodyList = []
    for i in range(1, maxRow):
        bodyCells = sheetData.cell(row = i, column = 4)
        bodyList.append(bodyCells.value)
    #checks to see if email is already in the sheet so theres no repeats
    if dic["body"] in bodyList:
        return True
    return False
